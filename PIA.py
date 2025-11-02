import datetime as dt
from tabulate import tabulate 
import json
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import sys
import sqlite3
from sqlite3 import Error
import warnings

warnings.filterwarnings('ignore', category=DeprecationWarning)

fecha_hoy = dt.date.today()
fecha_hoy = fecha_hoy.strftime("%m/%d/%Y")
fecha_hoy = dt.datetime.strptime(fecha_hoy, "%m/%d/%Y").date()
try:
    with sqlite3.connect("Eventos.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS CLIENTES (ID_CLIENTE INTEGER PRIMARY KEY, NOMBRE TEXT NOT NULL, APELLIDO TEXT NOT NULL);")
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS SALAS (ID_SALA INTEGER PRIMARY KEY, NOMBRE TEXT NOT NULL, CAPACIDAD INTEGER NOT NULL);")
        print("Se comenzo desde un estado vacio")
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS EVENTOS (ID_EVENTO INTEGER PRIMARY KEY, ID_SALA INTEGER, ID_CLIENTE INTEGER, NOMBRE_EVENTO TEXT NOT NULL, TURNO TEXT NOT NULL, FECHA timestamp, DISPONIBILIDAD INTEGER NOT NULL, FOREIGN KEY (ID_SALA) REFERENCES SALAS(ID_SALA), FOREIGN KEY (ID_CLIENTE) REFERENCES CLIENTES(ID_CLIENTE));")
except Error as e:
    print (e)
except:
    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

print("*"*70)
print("\t***Bienvenido al sistema de reservacion de eventos***")
print("*"*70)

while True:
    print("\n===============Menu===============")
    print("1.Registrar evento")
    print("2.Editar nombre del evento")
    print("3.Consultar reservaciones")
    print("4.Cancelar evento")
    print("5.Registrar cliente")
    print("6.Registrar sala")
    print("7.Salir")
    print("="*34)
    
    try:
        opcion = int(input("Ingrese una opcion: "))
    except ValueError:
        print("Favor de digitar un numero valido\n")
        continue
    
    match opcion:
        case 1:
            print("\n===============Registrar evento===============\n")
            try:
                with sqlite3.connect("Eventos.db") as conn:
                    mi_cursor = conn.cursor()
                    mi_cursor.execute("SELECT COUNT(*) FROM CLIENTES")
                    registros = mi_cursor.fetchall()
                    if registros[0][0] == 0:
                        print("Favor de registrar un cliente previamente para poder registrar un evento\n")
                        continue
                    mi_cursor.execute("SELECT COUNT(*) FROM SALAS")
                    registros = mi_cursor.fetchall()
                    if registros[0][0] == 0:
                        print("Favor de registrar una sala previamente para poder registrar un evento\n")
                        continue
            except Error as e:
                print (e)
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        
            lista_clientes = []
            claves_clientes_validas = []
            
            try:
                with sqlite3.connect("Eventos.db") as conn:
                    mi_cursor = conn.cursor()
                    mi_cursor.execute("SELECT ID_CLIENTE, APELLIDO, NOMBRE FROM CLIENTES ORDER BY APELLIDO, NOMBRE")
                    lista_clientes = mi_cursor.fetchall()
                    claves_clientes_validas = [cliente[0] for cliente in lista_clientes]
            except Error as e:
                print (e)
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                    
            while True:
                print("\n***************Clientes registrados***************")
                encabezados_clientes = ["ID Cliente", "Apellidos", "Nombre(s)"]
                print(tabulate(lista_clientes, headers=encabezados_clientes, tablefmt="fancy_grid", stralign="center", numalign="center"))
                
                try:
                    clave_cliente_elegida = int(input("Ingresa tu clave de cliente: "))                  
                    if clave_cliente_elegida not in claves_clientes_validas:
                        print("El cliente no existe\n")
                        salida = input("Escriba X si quiere regresar al menu principal, si no digita cualquier otro caracter: ")
                        if salida.upper() == "X":
                            break
                        continue
                except ValueError:
                    print("Favor de digitar un numero valido\n")
                    continue
                
                while True:     
                    fecha_elegida = input("\nIngrese la fecha del evento (mm/dd/aaaa): ")
                    try:
                        fecha_evento = dt.datetime.strptime(fecha_elegida, "%m/%d/%Y").date()
                    except ValueError:
                        print("Favor de digitar una fecha valida\n")
                        continue
                    if (fecha_evento - fecha_hoy).days <= 2:
                        print(f"La fecha debe ser, por lo menos, dos días posteriores a la fecha actual\n")
                        continue

                    if fecha_evento.weekday() == 6: 
                        lunes_siguiente = fecha_evento + dt.timedelta(days=1)                      
                        print(f"No se pueden realizar reservaciones para los dias domingo")
                        
                        opcion_domingo = input(f"Se propone reservar para el lunes siguiente ({lunes_siguiente.strftime('%m/%d/%Y')}), colocar S para aceptar: ")                            
                        if opcion_domingo.upper() == "S":
                            fecha_evento = lunes_siguiente
                            break 
                        else:
                            continue
                    break 
                    
                salas_info = {} 
                ids_salas_validas = []
                valor = (fecha_evento,) 
                
                try:
                    with sqlite3.connect("Eventos.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("SELECT ID_SALA, NOMBRE, CAPACIDAD FROM SALAS")
                        lista_salas_data = mi_cursor.fetchall() 
                        mi_cursor.execute("SELECT ID_SALA, TURNO FROM EVENTOS WHERE DATE(FECHA) = ? AND DISPONIBILIDAD = 1", valor)
                        eventos_en_fecha = mi_cursor.fetchall()
                except Error as e:
                    print(e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                    
                turnos_ocupados_por_sala = {}

                for sala_id, turno_ocupado in eventos_en_fecha:
                    if sala_id in turnos_ocupados_por_sala:
                        turnos_ocupados_por_sala[sala_id].append(turno_ocupado)
                    else:
                        turnos_ocupados_por_sala[sala_id] = [turno_ocupado]
                
                salas_info = {sala[0]: [sala[1], sala[2]] for sala in lista_salas_data}
                ids_salas_validas = list(salas_info.keys())

                turnos_validos = ["MATUTINO", "VESPERTINO", "NOCTURNO"]
                salas_turnos_disponibles = {} 
                filas_tabla_salas = []

                for sala_id, (nombre_sala, capacidad) in salas_info.items():
                    turnos_ocupados = turnos_ocupados_por_sala.get(sala_id, [])              
                    turnos_disponibles = [turno for turno in turnos_validos if turno not in turnos_ocupados]
                    salas_turnos_disponibles[sala_id] = [capacidad, turnos_disponibles] 
                    turnos = ', '.join(turnos_disponibles) if turnos_disponibles else "--- NINGUNO ---"
                    filas_tabla_salas.append([sala_id, nombre_sala, capacidad, turnos])

                print(f"\n\t*****Salas disponibles para la fecha {fecha_evento.strftime('%m/%d/%Y')}*****")
                filas_tabla_salas = []
                for sala_id, info in salas_turnos_disponibles.items():
                    nombre_sala = salas_info[sala_id][0]
                    capacidad = info[0]
                    turnos_str = ', '.join(info[1]) if info[1] else "--- NINGUNO ---"
                    filas_tabla_salas.append([sala_id, nombre_sala, capacidad, turnos_str])
                
                headers = ["Sala ID","Nombre Sala", "Cupo","Turnos Disponibles"]
                tabla = tabulate(filas_tabla_salas, headers, tablefmt="fancy_grid", stralign="center", numalign="center")
                print(tabla)
                
                while True:
                    try:
                        sala_elegida = int(input("Ingrese el ID de la sala: "))
                    except:
                        print("Favor de digitar un numero valido\n")
                        continue  
                    if sala_elegida not in ids_salas_validas:
                        print("La sala no existe\n")
                        continue
                    break 

                while True:
                    turno_elegido = input("Ingrese el turno a elegir: ").upper()
                    if turno_elegido not in turnos_validos:
                        print("Turno no valido \n")
                        continue
                    salida = ""
                    if turno_elegido not in salas_turnos_disponibles[sala_elegida][1]:
                        print("Este turno ya está ocupado para la sala y fecha seleccionadas. Por favor, elija otro turno\n")
                        salida = input("Escriba X si quiere regresar al menu principal, si no digita cualquier otro caracter: ")
                        if salida.upper() == "X":
                            break
                        continue
                    break
                if salida.upper() == "X":
                    break
                
                while True:
                    nombre_evento = input("Ingrese el nombre del evento: ")
                    if not nombre_evento:
                        print("El nombre del evento no puede estar vacio\n") 
                        continue
                    if nombre_evento.isspace():
                        print("El nombre del evento no puede consistir solo en espacios en blanco\n")
                        continue
                    if nombre_evento.isdigit():
                        print("El nombre del evento no puede ser un numero\n")
                        continue
                    break 
                
                valores = (sala_elegida, clave_cliente_elegida, nombre_evento.upper(), turno_elegido.upper(), fecha_evento, 1)

                try:
                    with sqlite3.connect("Eventos.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("INSERT INTO EVENTOS (ID_SALA, ID_CLIENTE, NOMBRE_EVENTO, TURNO, FECHA, DISPONIBILIDAD) VALUES (?, ?, ?, ?, ?,?)", valores)
                        mi_cursor.execute("PRAGMA foreign_keys = 1")
                        print("\n***Evento registrado con exito***")
                except Error as e:
                    print (e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                break 
        
        case 2:
            print("\n===============Editar nombre del evento===============\n")
            
            while True:
                fecha_inicio = input("Ingrese desde que fecha consultar los eventos (mm/dd/aaaa): ")
                try:
                    fecha_inicio = dt.datetime.strptime(fecha_inicio, "%m/%d/%Y").date()
                    fecha_inicio_iso = fecha_inicio.isoformat()
                    break
                except ValueError:
                    print("Favor de digitar una fecha valida\n")
                    continue

            while True:
                fecha_fin = input("Ingrese hasta que fecha consultar los eventos (mm/dd/aaaa): ")
                try:
                    fecha_fin = dt.datetime.strptime(fecha_fin, "%m/%d/%Y").date()
                    if fecha_fin < fecha_inicio:
                        print("La fecha final no puede ser menor a la fecha inicial\n")
                        continue
                    fecha_fin_iso = fecha_fin.isoformat()
                    break
                except ValueError:
                    print("Favor de digitar una fecha valida\n")
                    continue

            eventos_en_rango = []
            folios_eventos_validos = []
            
            try:
                with sqlite3.connect("Eventos.db", detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                    mi_cursor = conn.cursor()
                    consulta = """
                        SELECT E.ID_EVENTO, S.NOMBRE, C.NOMBRE, C.APELLIDO, E.NOMBRE_EVENTO, E.TURNO, strftime('%m/%d/%Y', E.FECHA)
                        FROM EVENTOS AS E INNER JOIN SALAS AS S ON E.ID_SALA = S.ID_SALA
                        INNER JOIN CLIENTES AS C ON E.ID_CLIENTE = C.ID_CLIENTE
                        WHERE DATE(E.FECHA) BETWEEN ? AND ? AND E.DISPONIBILIDAD = 1
                    """
                    valores = (fecha_inicio, fecha_fin)
                    mi_cursor.execute(consulta, valores)
                    eventos_en_rango = mi_cursor.fetchall() 
                    folios_eventos_validos = [evento[0] for evento in eventos_en_rango]
            except Error as e:
                print(f"Error de base de datos al consultar eventos: {e}")
                continue
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                continue

            if not eventos_en_rango:
                print(f"\nNo hay eventos registrados entre {fecha_inicio} y {fecha_fin}\n")
                continue
            else:
                print(f"\n\t**********Eventos registrados entre {fecha_inicio} y {fecha_fin}**********")
                
                filas_tabla_eventos = []
                for evento in eventos_en_rango:
                    folio = evento[0]
                    nombre_sala = evento[1]
                    nombre_cliente = f"{evento[2]} {evento[3]}"
                    nombre_evento = evento[4]
                    turno_evento = evento[5]
                    fecha_evento = evento[6]
                    
                    filas_tabla_eventos.append([folio, nombre_sala, nombre_cliente, nombre_evento, turno_evento, fecha_evento])
                
                headers = ["Folio del evento", "Sala", "Cliente", "Evento", "Turno", "Fecha"]
                tabla = tabulate(filas_tabla_eventos, headers, tablefmt="fancy_grid", stralign="center", numalign="center")
                print(tabla)

            while True:
                try:
                    folio_evento_elegido = int(input("Ingrese el folio del evento a editar: "))
                    if folio_evento_elegido not in folios_eventos_validos:
                        print("Elegir folio de evento dentro de las opciones mostradas\n")
                        print(tabla)
                        continue
                    break
                except ValueError:
                    print("Favor de digitar un numero valido\n")
                    continue
                    
            while True:
                nuevo_nombre_evento = input("Ingrese el nuevo nombre del evento: ")
                if not nuevo_nombre_evento:
                    print("El nombre del evento no puede estar vacio\n")
                    continue
                if nuevo_nombre_evento.isspace():
                    print("El nombre del evento no puede consistir solo en espacios en blanco\n")
                    continue
                if nuevo_nombre_evento.isdigit():
                    print("El nombre del evento no puede ser un numero\n")
                    continue
                nuevo_nombre = nuevo_nombre_evento.upper()
                valores = (nuevo_nombre, folio_evento_elegido)
                
                try:
                    with sqlite3.connect("Eventos.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("UPDATE EVENTOS SET NOMBRE_EVENTO = ? WHERE ID_EVENTO = ?", valores)
                        print("***Nombre del evento editado con exito***")
                        break
                except Error as e:
                    print(f"Error de base de datos al actualizar el evento: {e}")
                    break
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                    break
        case 3:
            print("\n===============Consultar reservaciones===============\n")

            while True:
                fecha_consultada = input(f"Ingrese la fecha a consultar (mm/dd/aaaa) o dejar vacio para asignar la fecha de hoy: ")
                
                if not fecha_consultada:
                    fecha_consulta_dt = fecha_hoy
                    break
                    
                try:
                    fecha_consulta_dt = dt.datetime.strptime(fecha_consultada, "%m/%d/%Y").date()
                    break
                except ValueError:
                    print("Favor de digitar una fecha valida\n")
                    continue
        
            filas_tabla = []
            valor = (fecha_consulta_dt,)    
            try:
                with sqlite3.connect("Eventos.db", detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                    mi_cursor = conn.cursor()
                    consulta = """
                        SELECT S.NOMBRE, C.NOMBRE, C.APELLIDO, E.NOMBRE_EVENTO, E.TURNO
                        FROM EVENTOS AS E INNER JOIN SALAS AS S ON E.ID_SALA = S.ID_SALA
                        INNER JOIN CLIENTES AS C ON E.ID_CLIENTE = C.ID_CLIENTE
                        WHERE DATE(E.FECHA) = ? AND E.DISPONIBILIDAD = 1
                    """   
                    mi_cursor.execute(consulta, valor)            
                    eventos_en_fecha = mi_cursor.fetchall()
            except Error as e:
                print(f"Error de base de datos al consultar eventos: {e}")
                continue
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                continue
            
            if not eventos_en_fecha:
                print(f"\nNo hay eventos registrados para la fecha {fecha_consulta_dt.strftime('%m/%d/%Y')}\n")
                continue
            else:
                for evento in eventos_en_fecha:
                    nombre_sala = evento[0]
                    nombre_cliente = f"{evento[1]} {evento[2]}"
                    nombre_evento = evento[3]
                    turno = evento[4]
                    
                    filas_tabla.append([nombre_sala, nombre_cliente, nombre_evento, turno])
                        
                headers = ["SALA","CLIENTE","EVENTO","TURNO"]

                tabla = tabulate(filas_tabla, headers, tablefmt="fancy_grid", stralign="center", numalign="center")

                print("*" * 70)
                print(f"\t** REPORTE DE RESERVACIONES PARA LA FECHA {fecha_consultada} **")
                print("*" * 70)
                print(tabla)
                print("*" * 70)
                print("\t**\t\tFIN DEL REPORTE\t\t**")    
                print("*" * 70)
                
                while True:
                    print("\n==========Menu de exportacion de datos==========")
                    print("A. Guardar en JSON")
                    print("B. Guardar en CSV")
                    print("C. Guardar en Excel")
                    opcion_exportacion = input("Ingrese una opcion o bien digite alguna otra tecla para regresar al menu principal: ")
                    
                    datos_a_exportar = [headers] + filas_tabla
                    nombre_base_archivo = "reservaciones"
                    
                    match opcion_exportacion.upper():
                        case "A":
                            datos_json = {
                                "fecha_consulta": fecha_consultada, 
                                "reservaciones": [{"sala": fila[0],
                                                "cliente": fila[1],
                                                "evento": fila[2],
                                                "turno": fila[3]
                                } for fila in filas_tabla]
                            }
                            
                            try:
                                with open("reservaciones.json", "w") as archivo:
                                    json.dump(datos_json, archivo, indent=4)
                                print("Datos guardados correctamente en formato JSON")
                            except Exception as e:
                                print(f"Error al guardar en JSON: {e}")

                        case "B":
                            try:
                                with open("reservaciones.csv","w", encoding="latin1", newline="") as archivo:
                                    grabador=csv.writer(archivo)
                                    grabador.writerows(datos_a_exportar)
                                print("Datos guardados correctamente en formato CSV")
                            except Exception as e:
                                print(f"Error al guardar en CSV: {e}")

                        case "C":
                            try:
                                libro = openpyxl.Workbook()
                                hoja = libro.active
                                hoja.title = "Reservaciones"
                                negritas = Font(bold=True)
                                borde_inferior = Border(bottom=Side(border_style="thick")) 
                                centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)

                                for col_num, header in enumerate(headers, start=1):
                                    celda = hoja.cell(row=1, column=col_num, value=header)
                                    celda.font = negritas
                                    celda.border = borde_inferior
                                    celda.alignment = centrado
                                
                                for fila_num, fila in enumerate(filas_tabla, start=2):
                                    for col_num, valor in enumerate(fila, start=1):
                                        celda = hoja.cell(row=fila_num, column=col_num, value=valor)
                                        celda.alignment = centrado

                                for col in ["A","B","C","D"]:
                                    hoja.column_dimensions[col].width = 30 

                                libro.save(f"reservaciones.xlsx") 
                                print("ARCHIVO EXCEL GENERADO")
                                print("Datos guardados correctamente en formato Excel")
                            except Exception as e:
                                print(f"Error al guardar en Excel: {e}")

                        case _:
                            break
                    break
        
        case 4:
            print("\n===============Cancelar evento===============\n")
            
            while True:
                fecha_inicio = input("Ingrese desde que fecha consultar los eventos (mm/dd/aaaa): ")
                try:
                    fecha_inicio = dt.datetime.strptime(fecha_inicio, "%m/%d/%Y").date()
                    fecha_inicio_iso = fecha_inicio.isoformat()
                    if (fecha_inicio - fecha_hoy).days <= 2:
                        print(f"La cancelacion solo puede realizarse con dos dias de anticipacion\n")
                        continue
                    break
                except ValueError:
                    print("Favor de digitar una fecha valida\n")
                    continue
                

            while True:
                fecha_fin = input("Ingrese hasta que fecha consultar los eventos (mm/dd/aaaa): ")
                try:
                    fecha_fin = dt.datetime.strptime(fecha_fin, "%m/%d/%Y").date()
                    if fecha_fin < fecha_inicio:
                        print("La fecha final no puede ser menor a la fecha inicial\n")
                        continue
                    fecha_fin_iso = fecha_fin.isoformat()
                    break
                except ValueError:
                    print("Favor de digitar una fecha valida\n")
                    continue
            
            eventos_en_rango = []
            folios_eventos_validos = []
            valores = (fecha_inicio, fecha_fin)
            
            try:
                with sqlite3.connect("Eventos.db", detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                    mi_cursor = conn.cursor()
                    consulta = """
                        SELECT ID_EVENTO, NOMBRE_EVENTO, TURNO, strftime('%m/%d/%Y', FECHA)
                        FROM EVENTOS 
                        WHERE DATE(FECHA) BETWEEN ? AND ? AND DISPONIBILIDAD = 1
                    """   
                    mi_cursor.execute(consulta, valores)
                    eventos_en_rango = mi_cursor.fetchall() 
                    folios_eventos_validos = [evento[0] for evento in eventos_en_rango]
            except Error as e:
                print(f"Error de base de datos al consultar eventos: {e}")
                continue
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                continue

            if not eventos_en_rango:
                print(f"\nNo hay eventos registrados entre {fecha_inicio.strftime('%m/%d/%Y')} y {fecha_fin.strftime('%m/%d/%Y')}\n")
                continue
            else:
                print(f"\n\t**********Eventos registrados entre {fecha_inicio.strftime('%m/%d/%Y')} y {fecha_fin.strftime('%m/%d/%Y')}**********")
                
                filas_tabla_eventos = []
                for evento in eventos_en_rango:
                    folio = evento[0]
                    nombre_evento = evento[1]
                    turno_evento = evento[2]
                    fecha_evento = evento[3]
                    
                    filas_tabla_eventos.append([folio, nombre_evento, turno_evento, fecha_evento])
                
                headers = ["Folio del evento", "Nombre delEvento", "Turno", "Fecha"]
                tabla = tabulate(filas_tabla_eventos, headers, tablefmt="fancy_grid", stralign="center", numalign="center")
                print(tabla)
            
            while True:
                try:
                    folio_evento_elegido = int(input("Ingrese el folio del evento a cancelar: "))
                    if folio_evento_elegido not in folios_eventos_validos:
                        print("Elegir folio de evento dentro de las opciones mostradas\n")
                        print(tabla)
                        continue
                    break
                except ValueError:
                    print("Favor de digitar un numero valido\n")
                    continue
            valor = (folio_evento_elegido,)
            
            while True:
                confirmacion = input("Colocar S para confirmar la cancelacion del evento, digitar cualquier otra tecla para regresar al menu principal: ")
                if confirmacion.upper() == "S":
                    try:
                        with sqlite3.connect("Eventos.db") as conn:
                            mi_cursor = conn.cursor()
                            mi_cursor.execute("UPDATE EVENTOS SET DISPONIBILIDAD = 0 WHERE ID_EVENTO = ?", (valor))
                            mi_cursor.execute("PRAGMA foreign_keys = 1")
                            print("***Evento cancelado con éxito***")
                            break
                    except Error as e:
                        print(f"Error de base de datos al cancelar el evento: {e}")
                        break
                    except:
                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        break
                else:
                    print("Operación cancelada por el usuario")
                break
        case 5:
            print("\n===============Registrar un cliente===============\n")
            while True:
                nombre = input("Ingrese nombre(s) del cliente: ")
                if not nombre:
                    print("El nombre no puede estar vacio\n") 
                    continue
                elif nombre.isdigit():
                    print("El nombre no puede ser un numero\n")    
                    continue
                elif nombre.isspace():
                    print("El nombre no puede consistir solo en espacios en blanco\n")
                    continue
                else:
                    break
                
            while True:
                apellido = input("Ingrese los apellidos del cliente: ")
                if not apellido:
                    print("El apellido no puede estar vacio\n") 
                    continue
                elif apellido.isdigit():
                    print("El apellido no puede ser un numero\n")
                    continue
                if apellido.isspace():
                    print("El apellido no puede consistir solo en espacios en blanco\n")
                    continue
                valores = (nombre.upper(), apellido.upper())
                try:
                    with sqlite3.connect("Eventos.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("SELECT NOMBRE FROM CLIENTES WHERE NOMBRE = ? AND APELLIDO = ?", valores)
                        registros = mi_cursor.fetchall()
                        if registros:
                            print("Ese cliente ya esta registrado\n")
                            continue
                except Error as e:
                    print (e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                
                try:
                    with sqlite3.connect("Eventos.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("INSERT INTO CLIENTES (NOMBRE, APELLIDO) VALUES(?,?)", valores)
                        mi_cursor.execute("PRAGMA foreign_keys = 1")
                        print("\n***Cliente registrado exitosamente***")
                except Error as e:
                    print (e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                break
          
        case 6:
            print("\n===============Registrar una sala===============\n")
            while True:
                nombre_sala = input("Ingrese el nombre de la sala: ")
                if not nombre_sala:
                    print("El nombre de la sala no puede estar vacio\n") 
                    continue
                if nombre_sala.isdigit():
                    print("El nombre de la sala no puede ser un numero\n")
                    continue
                if nombre_sala.isspace():
                    print("El nombre de la sala no puede consistir solo en espacios en blanco\n")
                    continue
                valor = (nombre_sala.upper(),)
                try:
                    with sqlite3.connect("Eventos.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("SELECT NOMBRE FROM SALAS WHERE NOMBRE = ?", valor)
                        registros = mi_cursor.fetchall()
                        if registros:
                            print("El nombre de la sala ya existe\n")
                            continue
                except Error as e:
                    print (e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                break
            
            while True:    
                try:
                    capacidad = int(input("Ingrese la capacidad de la sala: "))
                except ValueError:
                    print("Favor de digitar un numero valido\n")
                    continue
                if capacidad <= 0:
                    print("La capacidad debe ser mayor a 0")
                    continue
                valores = (nombre_sala.upper(), capacidad)
                try:
                    with sqlite3.connect("Eventos.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("INSERT INTO SALAS (NOMBRE, CAPACIDAD) VALUES(?,?)", valores)
                        mi_cursor.execute("PRAGMA foreign_keys = 1")
                        print("\n Sala registrada exitosamente")
                except Error as e:
                    print (e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                break
        
        case 7:
            decision = input("Escriba S para confirmar la salida del programa, para regresar al menu principal digite cualquier otra tecla: ")
            if decision.upper() == "S":
                print("*"*70)
                print("***Gracias por usar el sistema de reservacion, vuelva pronto***")
                print("*"*70)
                break
            continue
        
        case _:
            print("\nOpcion no valida, favor de intentarlo de nuevo")